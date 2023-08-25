import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup


# 打开网页
driver = webdriver.Chrome()  # 根据实际情况设置Chrome驱动程序的路径
driver.get('https://www.edbchinese.hk/lexlist_ch/')  # 替换为您要打开的网页的URL

# 读取 Excel 表格的内容
workbook = openpyxl.load_workbook('../excel/shengpizi.xlsx')  # 根据实际情况设置Excel文件的路径
worksheet1 = workbook['Sheet1']  # 替换为您要读取数据的工作表名称
worksheet2 = workbook['Sheet2']  # 替换为您要读取数据的工作表名称

# 打印 Excel 表格的内容
print('工作表名称：', worksheet1.title)
print(worksheet1.cell(row=2, column=4).value )
worksheets_first_row = 2  # 工作表中的第一行
for row in range(2, worksheet1.max_row + 1):
    row_order = worksheet1.cell(row=row, column=1).value  # 流水號
    teacher_order = worksheet1.cell(row=row, column=2).value  # 教育部字號
    key_unicode = worksheet1.cell(row=row, column=3).value  # unicode
    search_key = worksheet1.cell(row=row, column=4).value  # 假设要读取的数据在第一列
    
    if search_key is not None:
        search_page = driver.page_source
        search_soup = BeautifulSoup(search_page, 'html.parser')
        # 使用CSS选择器定位输入元素
        input_element = search_soup.select_one('div.search_text input[type=text]')

        # 检查输入元素是否存在
        if input_element:
            # 清空输入字段
            input_element['value'] = ''

            # 输入搜索关键字
            input_element['value'] = search_key
        else:
            print("未找到输入元素")


        # 使用CSS选择器定位搜索按钮元素
        search_button = search_soup.select_one('input.submit_btn')
        print(search_button)
        # 检查搜索按钮是否存在
        if search_button:
            # 点击搜索按钮
            # 点击搜索按钮
            search_button_parent = search_button.parent
            if search_button_parent.name == 'mainForm':
                form = search_button_parent
                form.submit()

            # TODO: 添加适当的等待代码
            wait = WebDriverWait(driver, 10)
            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'body > div.content > div.footer > a.top_btn')))
        else:
            print("未找到搜索按钮元素")

        # 等待网页加载完成

        # 获取网页源代码
        html = driver.page_source
    
        # 使用 BeautifulSoup 解析网页内容
        soup = BeautifulSoup(html, 'html.parser')
        
         # 根据实际情况定位结果元素
        simple_key = '' 
        # 获取多个数据body > div.content > div.content_text > div.content_right > table:nth-child(2) > tbody > tr:nth-child(2) > td:nth-child(1)
        
        # 获取<div class="content_right">
        content_right_div = soup.find('div', class_='content_right')

        # 获取第一个table的第二个tr的td
        jianhuazi = content_right_div.select('table:nth-of-type(1) tr:nth-of-type(2) td')[0].text.strip()

        # 获取第二个table的第二个tr的td
        bushou = content_right_div.select('table:nth-of-type(2) tr:nth-of-type(2) td:nth-of-type(1)')[0].text.strip()
        zongbihuashu = content_right_div.select('table:nth-of-type(2) tr:nth-of-type(2) td:nth-of-type(2)')[0].text.strip()
        # # 部首
        # bushou_elements = driver.find_elements_by_css_selector('body > div.content > div.content_text > div.content_right > table:nth-child(2) > tbody > tr:nth-child(2) > td:nth-child(1)')
        # bushou = bushou_elements[0].text.strip()
        # # 总笔画数
        # zongbihuashu_elements = driver.find_elements_by_css_selector('body > div.content > div.content_text > div.content_right > table:nth-child(2) > tbody > tr:nth-child(2) > td:nth-child(2)')
        # zongbihuashu = zongbihuashu_elements[0].text.strip()
        # 普通话
       # 定位并提取普通话发音列表
        pinyin_element = soup.find('td', class_='pinyin12')
        pinyin_text = pinyin_element.find('strong').text.strip()
        pinyin_list = [text.strip() for text in pinyin_text.split('\n') if text.strip()]

        # 定位并提取粤语发音列表
        jyutping_elements = soup.find_all('span', class_='jyutping12')
        jyutping_list = [element.find('strong').text.strip() for element in jyutping_elements]
        
        
        
        
        # 将网页的数据写回到新 Excel 表格,新建一行
        worksheet2.cell(row=worksheets_first_row, column=1).value = row_order  # 流水號
        worksheet2.cell(row=worksheets_first_row, column=2).value = teacher_order  # 流水號
        worksheet2.cell(row=worksheets_first_row, column=3).value = key_unicode  # 流水號
        worksheet2.cell(row=worksheets_first_row, column=4).value = search_key  # 流水號
        worksheet2.cell(row=worksheets_first_row, column=7).value =  simple_key # 簡化字
        worksheet2.cell(row=worksheets_first_row, column=8).value =  bushou # 簡化字
        worksheet2.cell(row=worksheets_first_row, column=9).value =  zongbihuashu # 簡化字
        for index, pinyin in enumerate(pinyin_list):
            worksheet2.cell(row=worksheets_first_row, column=10 + index).value = pinyin
        for index, jyutping in enumerate(jyutping_list):
            worksheet2.cell(row=worksheets_first_row, column=16 + index).value = jyutping    
        #换行
        worksheets_first_row += 1
        
        # 获取字词
        # 获取所有class为"ks1"或"ks2"的行
     # 查找包含"ks1"或"ks2"类的所有tr元素
        rows = soup.find_all('tr', class_=lambda value: value and ('ks1' in value or 'ks2' in value))

        # 遍历每一行
        for row in rows:
            # 获取字词
            word = row.select_one('.ci').text.strip()
            # 获取拼音
            word_pinyin = row.select_one('.pinyin').text.strip()
            # 获取粤语拼音
            word_jyutping = row.select_one('.jyutping').text.strip()
            # 获取阶段
            number = row.select('td')[-1].text.strip()
             # 每一个字詞一行
            worksheet2.cell(row=worksheets_first_row, column=1).value = row_order  # 流水號
            worksheet2.cell(row=worksheets_first_row, column=2).value = teacher_order  # 流水號
            worksheet2.cell(row=worksheets_first_row, column=3).value = key_unicode  # 流水號
            worksheet2.cell(row=worksheets_first_row,column=4).value = search_key  # 流水號
            worksheet2.cell(row=worksheets_first_row,column=5).value = word  # 流水號
            worksheet2.cell(row=worksheets_first_row,column=6).value = number  # 流水號
            
            worksheet2.cell(row=worksheets_first_row, column=10).value = word_pinyin
            worksheet2.cell(row=worksheets_first_row, column=16).value = word_jyutping
            worksheets_first_row += 1
             
        
        
    

# 保存并关闭 Excel 文件
workbook.save('../excel/shengpizi_result.xlsx')  # 根据实际情况设置Excel文件的路径
workbook.close()

# 关闭浏览器
driver.quit()