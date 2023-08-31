import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
import time
import re



import openpyxl

# 输入需要查询的字列表
workbook = openpyxl.load_workbook('../excel/duoyinzi.xlsx')
word_list_sheet = workbook['Sheet1'] 
# 创建保存表
word_detail_sheet = workbook['Sheet2']


driver = webdriver.Chrome()  # 根据实际情况设置Chrome驱动程序的路径


for row in range(2, word_list_sheet.max_row + 1):
    word = word_list_sheet.cell(row=row, column=1).value
    print(word)
    # 获取多音字组词页面
    driver.get('http://zuci.kaishicha.com/duoyinzizuci.asp') 
    wait = WebDriverWait(driver, 10)
    input_element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'form input[type=text]')))
    input_element.clear()
    input_element.send_keys(word)
        # 提交表单执行搜索
    form = driver.find_element(By.TAG_NAME, 'form')
    
    form.submit()
    driver.close()
    driver.switch_to.window(driver.window_handles[0])    
    # 获取新页面链接
    new_url = driver.current_url
    print(new_url)
    # 修改链接
    new_url = re.sub(r'/(\w+)\.html', r'/duoyin_\1.html', new_url)
    print(new_url)
    driver.get(new_url)
    wait = WebDriverWait(driver, 10)
    timeout = 30 # 超时秒数
    prev_count = 0
    while True:
        curr_count = driver.execute_script("return window.performance.getEntries().length;")
        
        if curr_count > prev_count:
            # 请求数增加,页面还在加载
            prev_count = curr_count
        else:
            # 请求数不再增加,页面加载完毕
            break
        time.sleep(1)
        
        timeout -= 1
        if timeout == 0:
            raise TimeoutException
        
    html = driver.page_source
    open('../html/duoyinzi_detail.html', 'w').write(html)
    soup = BeautifulSoup(html, 'html.parser')
    
    # 获取多音字列表
    duoyinzi_title_list = soup.find_all('span', class_='green')
    print(duoyinzi_title_list)
    for item in duoyinzi_title_list:
        title = item.text
        # 提取 "读" 和 "组词" 中间的字符
        du_pos = title.find('读') + 1  # 找到 "读" 字符的位置
        zu_pos = title.find('组词', du_pos) + 1  # 找到 "组词" 字符的位置
        duyin = title[du_pos:zu_pos-1]  # 提取中间的字符
        # 提取多音字列表,是当前item的下一个p标签节点
        cizu = item.find_next('p').find_all('a')
        cizu = [item.text for item in cizu]
        for cizu_item in cizu:
            word_detail_sheet.append([word, duyin, cizu_item])
            
    workbook.save('../excel/duoyinzi.xlsx')  
    open('./now_duoyinzi.txt','w').write(word)          
        
    

workbook.close()

# 关闭浏览器
driver.quit()