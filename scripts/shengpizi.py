import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
from selenium.common.exceptions import TimeoutException
import time

# 打开网页
driver = webdriver.Chrome()  # 根据实际情况设置Chrome驱动程序的路径

# 读取 Excel 表格的内容
workbook = openpyxl.load_workbook('../excel/shengpizi.xlsx')  # 根据实际情况设置Excel文件的路径
worksheet1 = workbook['Sheet1']  # 替换为您要读取数据的工作表名称
worksheet2 = workbook['Sheet2']  #

# 从文件读取开始行
start_row = int(open('./start_row.txt').read())

# 打印 Excel 表格的内容
worksheets_first_row = 2  # 保存表工作表中的第一行

# 获取最大行数和列数
max_row = worksheet2.max_row
max_column = worksheet2.max_column

# 获取保存表的最新行
last_data_row = 0
for row in range(max_row, 0, -1):
    row_values = [worksheet2.cell(row=row, column=column).value for column in range(1, max_column+1)]
    if any(row_values):
        last_data_row = row
        break
worksheets_first_row = last_data_row + 1    

for row in range(start_row, worksheet1.max_row + 1):
    row_order = worksheet1.cell(row=row, column=1).value  # 流水號
    teacher_order = worksheet1.cell(row=row, column=2).value  # 教育部字號
    key_unicode = worksheet1.cell(row=row, column=3).value  # unicode
    search_key = worksheet1.cell(row=row, column=4).value  # 假设要读取的数据在第一列
    driver.get('https://www.edbchinese.hk/lexlist_ch/')  # 
    wait = WebDriverWait(driver, 10)
    input_element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.search_text input[type=text]')))
    # 输入搜索词
    #input_element = driver.find_element(By.CSS_SELECTOR, 'div.search_text input[type=text]')
    input_element.clear()
    input_element.send_keys(search_key)
    
    # 提交表单执行搜索
    form = driver.find_element(By.TAG_NAME, 'form')
    form.submit()
    
    # 等待结果加载出来
   
    
    
    timeout = 30 # 超时秒数

    prev_count = 0
    while True:
        curr_count = driver.execute_script("return window.performance.getEntries().length;")
        
        if curr_count > prev_count:
            # 请求数增加,页面还在加载
            prev_count = curr_count
        else:
            # 请求数不再增加,页面加载完毕
            break
        time.sleep(1)
        
        timeout -= 1
        if timeout == 0:
            raise TimeoutException
        
    iframe = driver.find_element(By.ID, 'mainFrame')
    driver.switch_to.frame(iframe)    
     # 获取页面源码
    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')

    # 获取笔画、部首等信息
    open('../html/iframe.html', 'w').write(html)
    content_right = soup.select_one('div.content_right')
    if content_right is None: 
        worksheet2.cell(row=worksheets_first_row, column=1).value = row_order  # 流水號
        worksheet2.cell(row=worksheets_first_row, column=2).value = teacher_order  # 流水號
        worksheet2.cell(row=worksheets_first_row, column=3).value = key_unicode  # 流水號
        worksheet2.cell(row=worksheets_first_row, column=4).value = search_key  # 流水號
        worksheet2.cell(row=worksheets_first_row, column=5).value = '没查找到数据'  # 流水號
        worksheets_first_row += 1
        
        continue
    p_element = content_right.select_one('p')
    if p_element is not None and p_element.text.strip() == '抱歉，找不到合乎搜尋條件的條目。':
        worksheet2.cell(row=worksheets_first_row, column=1).value = row_order  # 流水號
        worksheet2.cell(row=worksheets_first_row, column=2).value = teacher_order  # 流水號
        worksheet2.cell(row=worksheets_first_row, column=3).value = key_unicode  # 流水號
        worksheet2.cell(row=worksheets_first_row, column=4).value = search_key  # 流水號
        worksheet2.cell(row=worksheets_first_row, column=5).value = '没查找到数据'  # 流水號
        worksheets_first_row += 1
        continue

    jianhuazi = content_right.select_one('table:nth-of-type(1) tr:nth-of-type(2) td')
    if jianhuazi:
        jianhuazi = jianhuazi.text.strip()
    else:
        jianhuazi = ''
    print('jianhuazi')    
    print(jianhuazi)    
    bushou = content_right.select_one('table:nth-of-type(2) tr:nth-of-type(2) td:nth-of-type(1)').text.strip()    
    print('bushou')
    print(bushou)
    zongbihuashu = content_right.select_one('table:nth-of-type(2) tr:nth-of-type(2) td:nth-of-type(2)').text.strip()
    print('zongbihuashu')
    print(zongbihuashu)
    # 获取拼音

    # 获取粤语拼音
    jyutping = [strong.text.strip() for strong in soup.select('span.jyutping12 strong')]
    pinyin_element = soup.find('td', class_='pinyin12')
    if pinyin_element:
        pinyin_text = pinyin_element.find('strong').text.strip()
    else:
        pinyin_text = ''    
    pinyin_list = [text.strip() for text in pinyin_text.split('\n') if text.strip()]
    print('pinyin')
    print(pinyin_list)
    
    # 定位并提取粤语发音列表
    jyutping_elements = soup.find_all('span', class_='jyutping12')
    if jyutping_elements:
        jyutping_list = [element.find('strong').text.strip() for element in jyutping_elements]
    else:
        jyutping_list = []    
    print('jyutping')
    print(jyutping_list)
    
    
    
    
    # 将网页的数据写回到新 Excel 表格,新建一行
    worksheet2.cell(row=worksheets_first_row, column=1).value = row_order  # 流水號
    worksheet2.cell(row=worksheets_first_row, column=2).value = teacher_order  # 流水號
    worksheet2.cell(row=worksheets_first_row, column=3).value = key_unicode  # 流水號
    worksheet2.cell(row=worksheets_first_row, column=4).value = search_key  # 流水號
    worksheet2.cell(row=worksheets_first_row, column=7).value =  jianhuazi # 簡化字
    worksheet2.cell(row=worksheets_first_row, column=8).value =  bushou # 簡化字
    worksheet2.cell(row=worksheets_first_row, column=9).value =  zongbihuashu # 簡化字
    for index, pinyin in enumerate(pinyin_list):
        worksheet2.cell(row=worksheets_first_row, column=10 + index).value = pinyin
    for index, jyutping in enumerate(jyutping_list):
        worksheet2.cell(row=worksheets_first_row, column=16 + index).value = jyutping    
    #换行
    worksheets_first_row += 1
    
    # 获取字词
    # 获取所有class为"ks1"或"ks2"的行
    # 查找包含"ks1"或"ks2"类的所有tr元素
    ks_rows = soup.find_all('tr', class_=lambda value: value and ('ks1' in value or 'ks2' in value))

    # 遍历每一行
    for ks_row in ks_rows:
        # 获取字词
        word = ks_row.select_one('.ci')
        if word:
            word = word.text.strip()
        else:
            word = ''    
        # 获取拼音
        word_pinyin = ks_row.select_one('.pinyinGreen')
        if word_pinyin:
            word_pinyin = word_pinyin.text.strip()
        else:
            word_pinyin = ''    
        
        # 获取粤语拼音
        word_jyutping = ks_row.select_one('.jyutpingGreen')
        if word_jyutping:
            word_jyutping = word_jyutping.text.strip()
        else:
            word_jyutping = ''    
        # 获取阶段
        number = ks_row.select('td')[-1].text.strip()
            # 每一个字詞一行
        worksheet2.cell(row=worksheets_first_row, column=1).value = row_order  # 流水號
        worksheet2.cell(row=worksheets_first_row, column=2).value = teacher_order  # 流水號
        worksheet2.cell(row=worksheets_first_row, column=3).value = key_unicode  # 流水號
        worksheet2.cell(row=worksheets_first_row,column=4).value = search_key  # 流水號
        worksheet2.cell(row=worksheets_first_row,column=5).value = word  # 流水號
        worksheet2.cell(row=worksheets_first_row,column=6).value = number  # 流水號
        
        worksheet2.cell(row=worksheets_first_row, column=10).value = word_pinyin
        worksheet2.cell(row=worksheets_first_row, column=16).value = word_jyutping
        worksheets_first_row += 1
    
    open('./start_row.txt','w').write(str(row + 1))           
    workbook.save('../excel/shengpizi.xlsx')      
        
    

# 保存并关闭 Excel 文件
# 根据实际情况设置Excel文件的路径
workbook.close()

# 关闭浏览器
driver.quit()