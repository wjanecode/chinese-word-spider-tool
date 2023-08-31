import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
import time
import re
import openpyxl
from pypinyin import pinyin, Style


import re

#encoding:utf-8

# map (final) constanant+tone to tone+constanan
mapConstTone2ToneConst = {'n1':  '1n',
                          'n2':  '2n',
                          'n3':  '3n',
                          'n4':  '4n',
                          'ng1': '1ng',
                          'ng2': '2ng',
                          'ng3': '3ng',
                          'ng4': '4ng',
                          'r1':  '1r',
                          'r2':  '2r',
                          'r3':  '3r',
                          'r4':  '4r',
                          'ng1': 'n1g',
                          'ng2': 'n2g',
                          'ng3': 'n3g',
                          'ng4': 'n4g',
                          }

# map vowel+vowel+tone to vowel+tone+vowel
mapVowelVowelTone2VowelToneVowel = {'ai1': 'a1i',
                                    'ai2': 'a2i',
                                    'ai3': 'a3i',
                                    'ai4': 'a4i',
                                    'ao1': 'a1o',
                                    'ao2': 'a2o',
                                    'ao3': 'a3o',
                                    'ao4': 'a4o',
                                    'ei1': 'e1i',
                                    'ei2': 'e2i',
                                    'ei3': 'e3i',
                                    'ei4': 'e4i',
                                    'ou1': 'o1u',
                                    'ou2': 'o2u',
                                    'ou3': 'o3u',
                                    'ou4': 'o4u'}

# map vowel-number combination to unicode
mapVowelTone2Unicode = {'a1': 'ā',
                        'a2': 'á',
                        'a3': 'ǎ',
                        'a4': 'à',
                        'e1': 'ē',
                        'e2': 'é',
                        'e3': 'ě',
                        'e4': 'è',
                        'i1': 'ī',
                        'i2': 'í',
                        'i3': 'ǐ',
                        'i4': 'ì',
                        'o1': 'ō',
                        'o2': 'ó',
                        'o3': 'ǒ',
                        'o4': 'ò',
                        'u1': 'ū',
                        'u2': 'ú',
                        'u3': 'ǔ',
                        'u4': 'ù',
                        'v1': 'ǜ',
                        'v2': 'ǘ',
                        'v3': 'ǚ',
                        'v4': 'ǜ',
                       }

def ConvertToneNumbersPinyin(lineIn):
    assert type(lineIn) is str
    lineOut = lineIn

    # mapVowelTone2Unicode
    for x, y in mapVowelTone2Unicode.items():
      lineOut = lineOut.replace(y, x).replace(y.upper(), x.upper())
    # mapVowelVowelTone2VowelToneVowel
    for x, y in mapVowelVowelTone2VowelToneVowel.items():
      lineOut = lineOut.replace(y, x).replace(y.upper(), x.upper())

    # first transform
    for x, y in mapConstTone2ToneConst.items():
      lineOut = lineOut.replace(y, x).replace(y.upper(), x.upper())

    return lineOut.replace('Ü', 'V').replace('ü', 'v')


  

    # sys.stdout.write(lineOut)

# 输入需要查询的字列表
workbook = openpyxl.load_workbook('../excel/duoyinzi.xlsx')

# 创建保存表
word_detail_sheet = workbook['Sheet2']

for row in range(1, word_detail_sheet.max_row + 1):
    word = word_detail_sheet.cell(row=row, column=1).value
    fasheng = word_detail_sheet.cell(row=row, column=2).value
    cizhu = word_detail_sheet.cell(row=row, column=3).value
    
    tone = ConvertToneNumbersPinyin(fasheng)
    tone = '[=' + tone + ']'
    
    fasheng_cizu = ''
    for ci in cizhu:
        fasheng_cizu += ci
        if ci == word:
            fasheng_cizu += tone
    
    word_detail_sheet.cell(row=row, column=5).value = fasheng_cizu  
workbook.save('../excel/duoyinzi.xlsx')    
    
